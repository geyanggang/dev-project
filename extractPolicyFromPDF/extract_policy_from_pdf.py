import os
import re
import pandas as pd
import pdfplumber
import glob
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import warnings

warnings.filterwarnings('ignore')


class InsuranceDataExtractor:
    def __init__(self, template_file):
        self.template_file = template_file
        # 读取模板文件获取准确的列结构
        self.template_df = pd.read_excel(template_file, sheet_name='Sheet1')
        self.required_columns = self.template_df.columns.tolist()
        print(f"模板列数: {len(self.required_columns)}")
        print(f"模板列名: {self.required_columns}")

        # 创建精确的列位置映射
        self.column_positions = self.create_exact_column_positions()

    def create_exact_column_positions(self):
        """创建精确的列位置映射"""
        positions = {
            # 基本信息 - 根据实际Excel结构调整
            '序号': 0,
            '名称': 1,
            '地址': 2,
            '证件类型': 3,
            '证件号码': 4,
            '联系电话': 5,
            '车辆信息': 6,
            '号牌号码': 7,
            '机动车种类': 8,
            '使用性质': 9,
            '发动机号': 10,
            '车辆识别代码': 11,
            '厂牌型号': 12,
            '功率': 13,
            '排量': 14,
            '登记日期': 15,
            '能源（燃料）种类': 16,
            '核定载客': 17,
            '核定载质量': 18,
            '保险单号': 19,

            # 商业险信息
            '新能源汽车损失保险保额（元）': 20,
            '新能源汽车损失保险费小计（元）': 21,
            '新能源汽车第三者责任保险保额（元）': 22,
            '新能源汽车第三者责任保险费小计（元）': 23,
            '车上人员责任险（司机）保额（元）': 24,
            '车上人员责任险（司机）保险费小计（元）': 25,
            '车上人员责任险（乘客）保额（元）': 26,
            '车上人员责任险（乘客）保险费小计（元）': 27,
            '保险费合计（元）': 28,
            '保险期限起期': 29,
            '保险期限起期止': 30,

            # 交强险信息
            '交强险保险单号': 31,
            '交强险保险费合计': 32,
            '死亡伤残赔偿限额': 33,
            '无责任死亡伤残赔偿限额': 34,
            '医疗费用赔偿限额': 35,
            '无责任医疗费用赔偿限额': 36,
            '财产损失赔偿限额': 37,
            '无责任财产损失赔偿限额': 38,
            '交强险保险期限起期': 39,
            '交强险保险期限起期止': 40
        }
        return positions

    def extract_plate_from_filename(self, filename):
        """从文件名中提取车牌号"""
        match = re.search(r'\((.*?)\)', filename)
        if match:
            return match.group(1)
        return ""

    def extract_commercial_insurance(self, pdf_path):
        """提取商业险信息 - 基于表格结构精确提取"""
        data = {}

        try:
            with pdfplumber.open(pdf_path) as pdf:
                # 处理第一页（商业险页面）
                page = pdf.pages[0]
                tables = page.extract_tables()

                if tables and len(tables) > 0:
                    table = tables[0]  # 第一个表格

                    # 提取被保险人信息
                    self.extract_policyholder_info_from_table(table, data)

                    # 提取车辆信息
                    self.extract_vehicle_info_from_table(table, data)

                    # 提取保险金额信息
                    self.extract_insurance_amounts_from_table(table, data)

                # 提取其他文本信息
                text = page.extract_text()
                self.extract_text_info(text, data)

        except Exception as e:
            print(f"提取商业险信息时出错: {str(e)}")

        return data

    def extract_policyholder_info_from_table(self, table, data):
        """从表格中提取被保险人信息"""
        try:
            # 名称 - 第1行第2列
            if len(table) > 1 and table[1][2]:
                data['名称'] = table[1][2].strip()

            # 地址 - 第2行第2列
            if len(table) > 2 and table[2][2]:
                data['地址'] = table[2][2].strip()

            # 证件类型 - 第3行第2列
            if len(table) > 3 and table[3][2]:
                data['证件类型'] = table[3][2].strip()

            # 证件号码 - 第3行第7列
            if len(table) > 3 and table[3][7]:
                data['证件号码'] = table[3][7].strip()

            # 联系电话 - 第3行第10列
            if len(table) > 3 and table[3][10]:
                data['联系电话'] = table[3][10].strip()

        except Exception as e:
            print(f"提取被保险人信息时出错: {str(e)}")

    def extract_vehicle_info_from_table(self, table, data):
        """从表格中提取车辆信息"""
        try:
            # 号牌号码 - 第4行第2列
            if len(table) > 4 and table[4][2]:
                data['号牌号码'] = table[4][2].strip()

            # 机动车种类 - 第4行第7列
            if len(table) > 4 and table[4][7]:
                data['机动车种类'] = table[4][7].strip()

            # 使用性质 - 第4行第10列
            if len(table) > 4 and table[4][10]:
                data['使用性质'] = table[4][10].strip()

            # 发动机号 - 第5行第2列
            if len(table) > 5 and table[5][2]:
                data['发动机号'] = table[5][2].strip()

            # 车辆识别代码 - 第5行第7列
            if len(table) > 5 and table[5][7]:
                data['车辆识别代码'] = table[5][7].strip()

            # 厂牌型号 - 第5行第10列
            if len(table) > 5 and table[5][10]:
                data['厂牌型号'] = table[5][10].strip()

            # 功率 - 第6行第2列
            if len(table) > 6 and table[6][2] is not None:
                power = str(table[6][2]).strip()
                data['功率'] = power if power and power != 'None' else ""

            # 排量 - 第6行第7列
            if len(table) > 6 and table[6][7]:
                data['排量'] = table[6][7].strip()

            # 登记日期 - 第6行第10列
            if len(table) > 6 and table[6][10]:
                data['登记日期'] = table[6][10].strip()

            # 能源（燃料）种类 - 第7行第2列
            if len(table) > 7 and table[7][2] is not None:
                energy = str(table[7][2]).strip()
                data['能源（燃料）种类'] = energy if energy and energy != 'None' else ""

            # 核定载客 - 第7行第7列
            if len(table) > 7 and table[7][7]:
                passenger_match = re.search(r'核定载客\s*(\d+)人', table[7][7])
                if passenger_match:
                    data['核定载客'] = passenger_match.group(1)
                else:
                    # 直接提取数字
                    passenger_match = re.search(r'(\d+)', table[7][7])
                    if passenger_match:
                        data['核定载客'] = passenger_match.group(1)

            # 核定载质量 - 第7行第9列
            if len(table) > 7 and table[7][10]:
                # 直接提取数值，不需要正则匹配
                load_value = table[7][10].strip()
                # 确保格式正确
                if '千克' not in load_value and load_value:
                    data['核定载质量'] = f"{load_value}千克"
                else:
                    data['核定载质量'] = load_value

        except Exception as e:
            print(f"提取车辆信息时出错: {str(e)}")

    def extract_insurance_amounts_from_table(self, table, data):
        """从表格中提取保险金额信息"""
        try:
            # 新能源汽车损失保险 - 第9行
            if len(table) > 9:
                if table[9][6]:  # 第6列是保额
                    data['新能源汽车损失保险保额（元）'] = self.parse_number(table[9][6])
                if table[9][10]:  # 第10列是保费
                    data['新能源汽车损失保险费小计（元）'] = self.parse_number(table[9][10])

            # 新能源汽车第三者责任保险 - 第10行
            if len(table) > 10:
                if table[10][6]:
                    data['新能源汽车第三者责任保险保额（元）'] = self.parse_number(table[10][6])
                if table[10][10]:
                    data['新能源汽车第三者责任保险费小计（元）'] = self.parse_number(table[10][10])

            # 车上人员责任险(司机) - 第11行
            if len(table) > 11:
                if table[11][6]:
                    data['车上人员责任险（司机）保额（元）'] = self.parse_number(table[11][6])
                if table[11][10]:
                    data['车上人员责任险（司机）保险费小计（元）'] = self.parse_number(table[11][10])

            # 车上人员责任险(乘客) - 第12行
            if len(table) > 12:
                if table[12][6]:  # 第6列是保额（特殊格式）
                    data['车上人员责任险（乘客）保额（元）'] = table[12][6].strip()
                if table[12][10]:  # 第10列是保费
                    data['车上人员责任险（乘客）保险费小计（元）'] = self.parse_number(table[12][10])

        except Exception as e:
            print(f"提取保险金额信息时出错: {str(e)}")

    def parse_number(self, value):
        """解析数字"""
        if not value:
            return 0
        try:
            # 移除逗号并转换为浮点数
            value_str = str(value).replace(',', '').replace(' ', '').strip()
            return float(value_str)
        except:
            return 0

    def extract_text_info(self, text, data):
        """提取文本信息"""
        # 保险单号
        policy_match = re.search(r'保险单号[：:]\s*(\d{20})', text)
        if policy_match:
            data['保险单号'] = policy_match.group(1)

        # 保险费合计 - 使用更精确的匹配
        total_match = re.search(r'保险费合计[：:].*?RMB\s*([\d,]+\.?\d*)', text)
        if total_match:
            data['保险费合计（元）'] = float(total_match.group(1).replace(',', ''))
        else:
            # 备用匹配模式
            total_match2 = re.search(r'保险费合计[：:]\s*([\d,]+\.?\d*)', text)
            if total_match2:
                data['保险费合计（元）'] = float(total_match2.group(1).replace(',', ''))

        # 保险期限
        period_match = re.search(r'保险期间自\s*([^起]+?)\s*起至\s*([^止]+?)\s*止', text)
        if period_match:
            data['保险期限起期'] = period_match.group(1).strip()
            data['保险期限起期止'] = period_match.group(2).strip()

    def extract_compulsory_insurance(self, pdf_path):
        """提取交强险信息"""
        data = {}

        try:
            with pdfplumber.open(pdf_path) as pdf:
                # 处理第二页（交强险页面）
                if len(pdf.pages) > 1:
                    page = pdf.pages[1]
                    tables = page.extract_tables()

                    if tables and len(tables) > 0:
                        table = tables[0]  # 第一个表格

                        # 提取交强险基本信息
                        self.extract_compulsory_info_from_table(table, data)

                    # 提取文本信息
                    text = page.extract_text()
                    self.extract_compulsory_text_info(text, data)

        except Exception as e:
            print(f"提取交强险信息时出错: {str(e)}")

        return data

    def extract_compulsory_info_from_table(self, table, data):
        """从表格中提取交强险信息"""
        try:
            # 责任限额
            if len(table) > 7:
                # 死亡伤残赔偿限额 - 第7行第5列
                if table[7][5]:
                    data['死亡伤残赔偿限额'] = table[7][5].strip()
                # 无责任死亡伤残赔偿限额 - 第7行第10列
                if table[7][10]:
                    data['无责任死亡伤残赔偿限额'] = table[7][10].strip()

            if len(table) > 8:
                # 医疗费用赔偿限额 - 第8行第5列
                if table[8][5]:
                    data['医疗费用赔偿限额'] = table[8][5].strip()
                # 无责任医疗费用赔偿限额 - 第8行第10列
                if table[8][10]:
                    data['无责任医疗费用赔偿限额'] = table[8][10].strip()

            if len(table) > 9:
                # 财产损失赔偿限额 - 第9行第5列
                if table[9][5]:
                    data['财产损失赔偿限额'] = table[9][5].strip()
                # 无责任财产损失赔偿限额 - 第9行第10列
                if table[9][10]:
                    data['无责任财产损失赔偿限额'] = table[9][10].strip()

        except Exception as e:
            print(f"提取交强险表格信息时出错: {str(e)}")

    def extract_compulsory_text_info(self, text, data):
        """提取交强险文本信息"""
        # 保险单号
        policy_match = re.search(r'保险单号[：:]\s*(\d{20})', text)
        if policy_match:
            data['交强险保险单号'] = policy_match.group(1)

        # 保险费 - 使用更精确的匹配
        premium_match = re.search(r'保险费合计[：:].*?RMB\s*(\d{1,3}(?:,\d{3})*\.?\d*)', text)
        if premium_match:
            data['交强险保险费合计'] = float(premium_match.group(1).replace(',', ''))
        else:
            # 备用匹配
            premium_match2 = re.search(r'保险费合计[：:]\s*([\d,]+\.?\d*)', text)
            if premium_match2:
                data['交强险保险费合计'] = float(premium_match2.group(1).replace(',', ''))

        # 保险期限 - 使用更精确的匹配
        period_match = re.search(r'保险期间自\s*([^起]+?)\s*起至\s*([^止]+?)\s*止', text)
        if period_match:
            data['交强险保险期限起期'] = period_match.group(1).strip()
            data['交强险保险期限起期止'] = period_match.group(2).strip()

    def parse_pdf(self, pdf_file):
        """解析单个PDF文件"""
        print(f"正在解析文件: {os.path.basename(pdf_file)}")

        data = {}

        try:
            # 从文件名提取车牌号
            filename = os.path.basename(pdf_file)
            plate_number = self.extract_plate_from_filename(filename)
            data['车辆信息'] = plate_number

            # 提取商业险信息
            commercial_data = self.extract_commercial_insurance(pdf_file)
            data.update(commercial_data)

            # 提取交强险信息
            compulsory_data = self.extract_compulsory_insurance(pdf_file)
            data.update(compulsory_data)

            # 计算保险费合计（如果未提取到）
            if not data.get('保险费合计（元）') and all([
                data.get('新能源汽车损失保险费小计（元）'),
                data.get('新能源汽车第三者责任保险费小计（元）'),
                data.get('车上人员责任险（司机）保险费小计（元）'),
                data.get('车上人员责任险（乘客）保险费小计（元）')
            ]):
                data['保险费合计（元）'] = (
                        data['新能源汽车损失保险费小计（元）'] +
                        data['新能源汽车第三者责任保险费小计（元）'] +
                        data['车上人员责任险（司机）保险费小计（元）'] +
                        data['车上人员责任险（乘客）保险费小计（元）']
                )

            print(f"✓ 成功解析: {os.path.basename(pdf_file)}")
            return data

        except Exception as e:
            print(f"✗ 解析PDF文件 {os.path.basename(pdf_file)} 时出错: {str(e)}")
            return None

    def save_to_excel(self, data_list, output_file):
        """直接保存数据到Excel文件"""
        try:
            # 加载模板工作簿
            book = load_workbook(self.template_file)
            sheet = book.active

            # 找到数据开始的行（跳过表头）
            data_start_row = 2  # 假设数据从第2行开始

            # 清除模板中可能存在的旧数据（保留表头）
            for row in range(data_start_row, sheet.max_row + 1):
                for col in range(1, len(self.required_columns) + 1):
                    sheet.cell(row=row, column=col).value = None

            # 写入新数据
            for i, data in enumerate(data_list):
                row_num = data_start_row + i

                # 设置序号
                sheet.cell(row=row_num, column=1).value = i + 1

                # 按位置写入数据
                for field, col_num in self.column_positions.items():
                    if field in data and data[field] is not None:
                        value = data[field]

                        # 处理NaN值
                        if pd.isna(value):
                            value = ""

                        # 特殊处理保险单号等长数字，避免科学计数法
                        if field in ['保险单号', '交强险保险单号'] and value:
                            sheet.cell(row=row_num, column=col_num + 1).value = str(value)
                            sheet.cell(row=row_num, column=col_num + 1).number_format = '@'  # 文本格式
                        elif isinstance(value, float):
                            # 保留两位小数
                            sheet.cell(row=row_num, column=col_num + 1).value = round(value, 2)
                            sheet.cell(row=row_num, column=col_num + 1).number_format = '0.00'
                        else:
                            sheet.cell(row=row_num, column=col_num + 1).value = value

            # 保存文件
            book.save(output_file)
            print(f"✓ 数据已按照模板格式保存到: {output_file}")
            return True

        except Exception as e:
            print(f"✗ 保存文件时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def process_pdfs(self, pdf_folder, output_file):
        """处理所有PDF文件并更新Excel"""
        # 查找所有PDF文件
        pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))
        if not pdf_files:
            pdf_files = glob.glob(os.path.join(pdf_folder, "**", "*.pdf"), recursive=True)

        print(f"找到 {len(pdf_files)} 个PDF文件")

        if not pdf_files:
            print("错误: 没有找到任何PDF文件!")
            return None

        all_data = []
        success_count = 0

        for pdf_file in pdf_files:
            data = self.parse_pdf(pdf_file)
            if data:
                all_data.append(data)
                success_count += 1

        print(f"\n成功解析 {success_count}/{len(pdf_files)} 个文件")

        if not all_data:
            print("错误: 没有成功解析任何PDF文件")
            return None

        # 调试信息：显示实际写入的数据
        print("\n调试信息 - 即将写入Excel的数据:")
        for data in all_data:
            for field, value in data.items():
                if value and str(value).strip() and str(value) != 'nan':
                    print(f"  {field}: {value}")

        # 使用模板格式保存
        if self.save_to_excel(all_data, output_file):
            # 创建DataFrame用于返回
            new_df = pd.DataFrame(all_data)
            return new_df
        else:
            print("✗ 保存文件失败")
            return None


def main():
    print("=" * 50)
    print("   保险保单信息提取工具")
    print("=" * 50)

    # 获取当前脚本所在目录
    if getattr(sys, 'frozen', False):
        # 如果是打包后的exe文件
        current_dir = os.path.dirname(sys.executable)
    else:
        # 如果是Python脚本
        current_dir = os.path.dirname(os.path.abspath(__file__))

    # 设置文件路径
    template_file = os.path.join(current_dir, "模板.xlsx")
    pdf_folder = current_dir  # 在当前目录查找PDF
    output_file = os.path.join(current_dir, "保险数据汇总.xlsx")

    print(f"工作目录: {current_dir}")
    print(f"模板文件: {template_file}")
    print(f"输出文件: {output_file}")
    print("-" * 50)

    if not os.path.exists(template_file):
        print("错误: 未找到模板文件!")
        print("请确保 '模板.xlsx' 文件存在")
        input("按任意键退出...")
        return

    # 创建提取器实例
    extractor = InsuranceDataExtractor(template_file)

    # 处理PDF文件
    result_df = extractor.process_pdfs(pdf_folder, output_file)

    if result_df is not None:
        print(f"\n🎉 处理完成！共提取 {len(result_df)} 条保单信息")
        print(f"💾 文件已保存至: {output_file}")

        # 显示提取结果与模板的对比
        print("\n提取结果预览:")
        if result_df is not None and not result_df.empty:
            # 显示关键字段
            key_fields = ['名称', '证件类型', '联系电话', '号牌号码', '机动车种类', '使用性质',
                          '发动机号', '车辆识别代码', '厂牌型号', '功率', '排量', '登记日期',
                          '核定载客', '核定载质量', '保险单号', '新能源汽车损失保险保额（元）',
                          '新能源汽车损失保险费小计（元）', '新能源汽车第三者责任保险保额（元）',
                          '新能源汽车第三者责任保险费小计（元）', '车上人员责任险（司机）保额（元）',
                          '车上人员责任险（司机）保险费小计（元）', '车上人员责任险（乘客）保额（元）',
                          '车上人员责任险（乘客）保险费小计（元）', '保险费合计（元）',
                          '交强险保险单号', '交强险保险费合计', '保险期限起期', '保险期限起期止']

            for field in key_fields:
                if field in result_df.columns:
                    value = result_df.iloc[0][field]
                    if pd.isna(value) or value == "":
                        print(f"  {field}: 未提取到")
                    else:
                        print(f"  {field}: {value}")
    else:
        print("\n❌ 处理失败！")

    print("\n按任意键退出...")
    input()


if __name__ == "__main__":
    main()